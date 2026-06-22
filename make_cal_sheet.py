from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F3864")
SUB_FILL = PatternFill("solid", start_color="2E5496")
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")   # cells to fill in
META_FILL = PatternFill("solid", start_color="DDEBF7")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "LibreVNA Cal 915MHz"
ws.sheet_view.showGridLines = False

# Column widths
widths = [16, 20, 16, 14, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def style_block(cell, fill, color="000000", bold=True, size=11, align="left"):
    cell.font = Font(name=FONT, bold=bold, color=color, size=size)
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")

# Title
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "LibreVNA Output Power Calibration  —  915 MHz"
style_block(c, HDR_FILL, color="FFFFFF", size=14, align="center")
ws.row_dimensions[1].height = 26

# Metadata block
meta = [
    ("Date", "2026-06-17"),
    ("Operator", ""),
    ("Test frequency (MHz)", 915),
    ("Reference plane", "VNA cable end @ PCB connector"),
    ("Power meter", "HP 438A"),
    ("Sensor", "HP 8481A"),
    ("Ref cal factor @ 50 MHz (%)", 100.0),
    ("Meas cal factor @ 915 MHz (%)", 99.0),
    ("Cable / adapter notes", ""),
    ("Avg / settling notes", "increase averaging below -25 dBm; re-zero first"),
]
row = 3
ws.merge_cells(f"A{row}:E{row}")
style_block(ws[f"A{row}"], SUB_FILL, color="FFFFFF", align="left")
ws[f"A{row}"] = "Setup / Conditions"
row += 1
for label, val in meta:
    lc = ws[f"A{row}"]
    lc.value = label
    lc.font = Font(name=FONT, bold=True, size=10)
    lc.fill = META_FILL
    lc.alignment = Alignment(horizontal="left", vertical="center")
    lc.border = BORDER
    ws.merge_cells(f"B{row}:E{row}")
    vc = ws[f"B{row}"]
    vc.value = val
    vc.font = Font(name=FONT, size=10, color="0000FF")
    vc.fill = INPUT_FILL
    vc.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = BORDER
    row += 1

# Data table header
row += 1
table_hdr_row = row
headers = ["VNA set (dBm)", "HP 438A reading (dBm)", "Delta (dBm)", "Re-zeroed?", "Notes"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col, value=h)
    style_block(cell, SUB_FILL, color="FFFFFF", size=10, align="center")
    cell.border = BORDER
ws.row_dimensions[row].height = 30
for col in range(1, 6):
    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data rows: 0 down to -30 dBm, 1 dB steps
first_data = row + 1
for i, setpt in enumerate(range(0, -31, -1)):
    r = first_data + i
    # VNA set (input, but pre-filled per plan)
    a = ws.cell(row=r, column=1, value=setpt)
    a.font = Font(name=FONT, size=10, color="0000FF")
    a.alignment = Alignment(horizontal="center")
    a.fill = INPUT_FILL
    # reading (to be entered)
    b = ws.cell(row=r, column=2)
    b.font = Font(name=FONT, size=10, color="0000FF")
    b.alignment = Alignment(horizontal="center")
    b.fill = INPUT_FILL
    b.number_format = "0.00"
    # delta = reading - set (formula)
    d = ws.cell(row=r, column=3, value=f"=IF(B{r}=\"\",\"\",B{r}-A{r})")
    d.font = Font(name=FONT, size=10)
    d.alignment = Alignment(horizontal="center")
    d.number_format = "0.00"
    # re-zeroed
    e = ws.cell(row=r, column=4)
    e.fill = INPUT_FILL
    e.alignment = Alignment(horizontal="center")
    # notes
    f = ws.cell(row=r, column=5)
    f.font = Font(name=FONT, size=10, color="0000FF")
    f.fill = INPUT_FILL
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = BORDER
last_data = first_data + 30

# Summary stats
srow = last_data + 2
ws.cell(row=srow, column=1, value="Summary").font = Font(name=FONT, bold=True, size=11)
stats = [
    ("Mean delta (dBm)", f"=IFERROR(AVERAGE(C{first_data}:C{last_data}),\"\")"),
    ("Min delta (dBm)", f"=IFERROR(MIN(C{first_data}:C{last_data}),\"\")"),
    ("Max delta (dBm)", f"=IFERROR(MAX(C{first_data}:C{last_data}),\"\")"),
    ("Delta span (dBm)", f"=IFERROR(MAX(C{first_data}:C{last_data})-MIN(C{first_data}:C{last_data}),\"\")"),
    ("Points entered", f"=COUNT(B{first_data}:B{last_data})"),
]
for k, (label, formula) in enumerate(stats):
    r = srow + 1 + k
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10)
    lc.fill = META_FILL
    lc.border = BORDER
    vc = ws.cell(row=r, column=2, value=formula)
    vc.font = Font(name=FONT, size=10)
    vc.number_format = "0.00"
    vc.border = BORDER
    vc.alignment = Alignment(horizontal="center")

# Freeze panes below table header
ws.freeze_panes = f"A{first_data}"

wb.save("LibreVNA_cal_915MHz.xlsx")
print("saved", first_data, last_data)
