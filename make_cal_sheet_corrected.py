from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F3864")
SUB_FILL = PatternFill("solid", start_color="2E5496")
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")
CALC_FILL = PatternFill("solid", start_color="E2EFDA")
META_FILL = PatternFill("solid", start_color="DDEBF7")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

sets = list(range(0, -31, -1))
raw = [-3.27,-3.28,-3.44,-4.44,-5.29,-6.29,-7.5,-8.5,-9.65,-10.65,-11.54,-12.66,
       -13.68,-14.55,-15.56,-16.78,-17.8,-18.96,-19.98,-20.89,-21.94,-22.84,-23.91,
       -25.18,-26.27,-27.52,-28.7,-29.84,-31.16,-32.43,-33.98]

wb = Workbook()
ws = wb.active
ws.title = "LibreVNA Cal 908MHz"
ws.sheet_view.showGridLines = False

widths = [15, 17, 17, 13, 13, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def style(cell, fill, color="000000", bold=True, size=11, align="left"):
    cell.font = Font(name=FONT, bold=bold, color=color, size=size)
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")

ws.merge_cells("A1:F1")
style(ws["A1"], HDR_FILL, color="FFFFFF", size=14, align="center")
ws["A1"] = "LibreVNA Output Power Calibration  —  908 MHz  (cal-factor corrected)"
ws.row_dimensions[1].height = 26

meta = [
    ("Date", "2026-06-17"),
    ("Operator", ""),
    ("Test frequency (MHz)", 908),
    ("Reference plane", "VNA cable end @ PCB connector"),
    ("Power meter", "HP 438A"),
    ("Sensor", "HP 8481A"),
    ("Ref cal factor @ 50 MHz (%)", 100.0),
    ("Meas cal factor @ 908 MHz (%)", 98.85),   # <-- drives the correction
]
CF_ROW = None
row = 3
ws.merge_cells(f"A{row}:F{row}")
style(ws[f"A{row}"], SUB_FILL, color="FFFFFF")
ws[f"A{row}"] = "Setup / Conditions"
row += 1
for label, val in meta:
    lc = ws[f"A{row}"]; lc.value = label
    lc.font = Font(name=FONT, bold=True, size=10); lc.fill = META_FILL
    lc.alignment = Alignment(horizontal="left", vertical="center"); lc.border = BORDER
    ws.merge_cells(f"B{row}:F{row}")
    vc = ws[f"B{row}"]; vc.value = val
    vc.font = Font(name=FONT, size=10, color="0000FF"); vc.fill = INPUT_FILL
    vc.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BORDER
    if "Meas cal factor" in label:
        CF_ROW = row
    row += 1

CF = f"$B${CF_ROW}"

row += 1
headers = ["VNA set (dBm)", "Raw reading (dBm)", "CF-corrected (dBm)",
           "Delta (dBm)", "Cable S21 (dB)", "Notes"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col, value=h)
    style(cell, SUB_FILL, color="FFFFFF", size=10, align="center")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
ws.row_dimensions[row].height = 30

first = row + 1
for i, (s, rv) in enumerate(zip(sets, raw)):
    r = first + i
    a = ws.cell(row=r, column=1, value=s)
    a.font = Font(name=FONT, size=10, color="0000FF"); a.fill = INPUT_FILL
    a.alignment = Alignment(horizontal="center")
    b = ws.cell(row=r, column=2, value=rv)
    b.font = Font(name=FONT, size=10, color="0000FF"); b.fill = INPUT_FILL
    b.alignment = Alignment(horizontal="center"); b.number_format = "0.00"
    # corrected = raw - 10*log10(CF/100)
    c = ws.cell(row=r, column=3, value=f"=IF(B{r}=\"\",\"\",B{r}-10*LOG10({CF}/100))")
    c.font = Font(name=FONT, size=10); c.fill = CALC_FILL
    c.alignment = Alignment(horizontal="center"); c.number_format = "0.00"
    d = ws.cell(row=r, column=4, value=f"=IF(C{r}=\"\",\"\",C{r}-A{r})")
    d.font = Font(name=FONT, size=10)
    d.alignment = Alignment(horizontal="center"); d.number_format = "0.00"
    e = ws.cell(row=r, column=5, value=-0.02)
    e.font = Font(name=FONT, size=10, color="0000FF"); e.fill = INPUT_FILL
    e.alignment = Alignment(horizontal="center"); e.number_format = "0.00"
    f = ws.cell(row=r, column=6)
    f.font = Font(name=FONT, size=10, color="0000FF"); f.fill = INPUT_FILL
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = BORDER
last = first + len(sets) - 1

srow = last + 2
ws.cell(row=srow, column=1, value="Summary (corrected)").font = Font(name=FONT, bold=True, size=11)
stats = [
    ("Mean delta, usable band −2…−25 (dBm)", f"=IFERROR(AVERAGE(D{first+2}:D{first+25}),\"\")"),
    ("Min delta, full sweep (dBm)", f"=IFERROR(MIN(D{first}:D{last}),\"\")"),
    ("Max delta, full sweep (dBm)", f"=IFERROR(MAX(D{first}:D{last}),\"\")"),
    ("Points entered", f"=COUNT(B{first}:B{last})"),
]
for k, (label, formula) in enumerate(stats):
    r = srow + 1 + k
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10); lc.fill = META_FILL; lc.border = BORDER
    ws.merge_cells(f"A{r}:C{r}")
    for col in (1,2,3): ws.cell(row=r, column=col).border = BORDER
    vc = ws.cell(row=r, column=4, value=formula)
    vc.font = Font(name=FONT, size=10); vc.number_format = "0.00"
    vc.border = BORDER; vc.alignment = Alignment(horizontal="center")

ws.freeze_panes = f"A{first}"

out = r"C:\Users\Benjamin\Desktop\LibreVNA_cal_908MHz_corrected.xlsx"
wb.save(out)
print("saved", out, "CF cell", CF, "data", first, last)
