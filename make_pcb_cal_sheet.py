from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HDR_FILL  = PatternFill("solid", start_color="1F3864")
SUB_FILL  = PatternFill("solid", start_color="2E5496")
INPUT_FILL= PatternFill("solid", start_color="FFF2CC")
CALC_FILL = PatternFill("solid", start_color="E2EFDA")
META_FILL = PatternFill("solid", start_color="DDEBF7")
LOCK_FILL = PatternFill("solid", start_color="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

# True input power at PCB from corrected VNA cal table (-2 to -25 dBm)
# (CF-corrected column, cable S21 = -0.02 dB already absorbed)
data = [
    (-2,  -3.39),
    (-3,  -4.39),
    (-4,  -5.24),
    (-5,  -6.24),
    (-6,  -7.45),
    (-7,  -8.45),
    (-8,  -9.60),
    (-9,  -10.60),
    (-10, -11.49),
    (-11, -12.61),
    (-12, -13.63),
    (-13, -14.50),
    (-14, -15.51),
    (-15, -16.73),
    (-16, -17.75),
    (-17, -18.91),
    (-18, -19.93),
    (-19, -20.84),
    (-20, -21.89),
    (-21, -22.79),
    (-22, -23.86),
    (-23, -25.13),
    (-24, -26.22),
    (-25, -27.47),
]

wb = Workbook()
ws = wb.active
ws.title = "PCB Cal 908MHz"
ws.sheet_view.showGridLines = False

widths = [16, 20, 14, 12, 12, 24]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def style(cell, fill, color="000000", bold=True, size=11, align="left"):
    cell.font = Font(name=FONT, bold=bold, color=color, size=size)
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center")

# Title
ws.merge_cells("A1:F1")
style(ws["A1"], HDR_FILL, color="FFFFFF", size=14, align="center")
ws["A1"] = "AD8317 PCB Calibration  —  908 MHz  (vs calibrated VNA)"
ws.row_dimensions[1].height = 26

# Metadata
meta = [
    ("Date", "2026-06-17"),
    ("Operator", ""),
    ("Test frequency (MHz)", 908),
    ("Reference plane", "PCB SMA connector input"),
    ("Signal source", "LibreVNA (calibrated, corrected table)"),
    ("PCB ADC pin", "GPIO36 / ADC1_CH0, 12-bit, ATTEN_DB_0"),
    ("Onboard pad", "10 dB cleanup pad (included in true input values)"),
    ("Firmware", "rfmeter1.ino  —  read avg_raw from Serial @ 115200"),
    ("Serial column to read", "3rd column: avg_raw"),
]
row = 3
ws.merge_cells(f"A{row}:F{row}")
style(ws[f"A{row}"], SUB_FILL, color="FFFFFF")
ws[f"A{row}"] = "Setup"
row += 1
for label, val in meta:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10)
    lc.fill = META_FILL; lc.border = BDR
    lc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"B{row}:F{row}")
    vc = ws.cell(row=row, column=2, value=val)
    vc.font = Font(name=FONT, size=10, color="0000FF")
    vc.fill = INPUT_FILL
    vc.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 7):
        ws.cell(row=row, column=col).border = BDR
    row += 1

# Table header
row += 1
hdr_row = row
headers = ["VNA set (dBm)", "True input at PCB (dBm)", "avg_raw (enter)", "min_raw", "max_raw", "Notes"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col, value=h)
    style(cell, SUB_FILL, color="FFFFFF", size=10, align="center")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BDR
ws.row_dimensions[row].height = 30

first = row + 1
for i, (vna_set, true_in) in enumerate(data):
    r = first + i

    a = ws.cell(row=r, column=1, value=vna_set)
    a.font = Font(name=FONT, size=10); a.fill = LOCK_FILL
    a.alignment = Alignment(horizontal="center"); a.border = BDR
    a.number_format = "0"

    b = ws.cell(row=r, column=2, value=true_in)
    b.font = Font(name=FONT, size=10); b.fill = LOCK_FILL
    b.alignment = Alignment(horizontal="center"); b.border = BDR
    b.number_format = "0.00"

    c = ws.cell(row=r, column=3)
    c.font = Font(name=FONT, size=10, color="0000FF"); c.fill = INPUT_FILL
    c.alignment = Alignment(horizontal="center"); c.border = BDR

    d = ws.cell(row=r, column=4)
    d.font = Font(name=FONT, size=10, color="0000FF"); d.fill = INPUT_FILL
    d.alignment = Alignment(horizontal="center"); d.border = BDR

    e = ws.cell(row=r, column=5)
    e.font = Font(name=FONT, size=10, color="0000FF"); e.fill = INPUT_FILL
    e.alignment = Alignment(horizontal="center"); e.border = BDR

    f = ws.cell(row=r, column=6)
    f.font = Font(name=FONT, size=10, color="0000FF"); f.fill = INPUT_FILL
    f.border = BDR

last = first + len(data) - 1

# Summary
srow = last + 2
ws.cell(row=srow, column=1, value="Summary").font = Font(name=FONT, bold=True, size=11)
stats = [
    ("Points entered", f"=COUNT(C{first}:C{last})"),
    ("avg_raw at top (set −2 dBm)", f"=C{first}"),
    ("avg_raw at bottom (set −25 dBm)", f"=C{last}"),
    ("Total count span", f"=IFERROR(C{first}-C{last},\"\")"),
]
for k, (label, formula) in enumerate(stats):
    r = srow + 1 + k
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(name=FONT, bold=True, size=10)
    lc.fill = META_FILL; lc.border = BDR
    ws.merge_cells(f"A{r}:B{r}")
    for col in (1,2): ws.cell(row=r, column=col).border = BDR
    vc = ws.cell(row=r, column=3, value=formula)
    vc.font = Font(name=FONT, size=10)
    vc.border = BDR; vc.alignment = Alignment(horizontal="center")

ws.freeze_panes = f"A{first}"

out = r"C:\Users\Benjamin\Desktop\PCB_cal_908MHz.xlsx"
wb.save(out)
print("saved", out)
